"""
転記先（見積書）Excelシートへの書き込み処理。

転記先レイアウトは 3行標準 / 2行標準 / 1行標準 のプリセットから選ぶ。
数量・単位・単価は各ブロック1行目に書く（金額式が1行目を参照するため）。
金額列（IF式）には書き込まない。
"""

from copy import copy

from openpyxl.formula.translate import Translator
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from logger import Logger


LAYOUT_3_ROWS = '3行標準'
LAYOUT_2_ROWS = '2行標準'
LAYOUT_1_ROW = '1行標準'

BLOCK_ROWS_BY_LAYOUT = {
    LAYOUT_3_ROWS: 3,
    LAYOUT_2_ROWS: 2,
    LAYOUT_1_ROW: 1,
}


def get_block_rows(settings: dict) -> int:
    """設定された転記先レイアウトの1明細あたり行数を返す。"""
    layout = settings.get('dst_layout', LAYOUT_3_ROWS)
    return BLOCK_ROWS_BY_LAYOUT.get(layout, 3)


def _join_lines(*values) -> str:
    """空の要素を除外し、セル内改行で複数行情報をまとめる。"""
    return '\n'.join(str(v).strip() for v in values if str(v or '').strip())


def _col(letter: str) -> int:
    return column_index_from_string(letter.upper())


def _safe_set(ws, row: int, col_letter: str, value) -> None:
    """結合セル対応のセル書き込み。結合範囲なら左上セルに書く。"""
    col = _col(col_letter)
    for merge in ws.merged_cells.ranges:
        if (merge.min_row <= row <= merge.max_row
                and merge.min_col <= col <= merge.max_col):
            ws.cell(merge.min_row, merge.min_col).value = value
            return
    ws.cell(row, col).value = value


def _copy_cell(src_cell, dst_cell) -> None:
    """値・数式・書式をテンプレートセルからコピーする。"""
    value = src_cell.value
    if isinstance(value, str) and value.startswith('='):
        value = Translator(value, origin=src_cell.coordinate).translate_formula(
            dst_cell.coordinate)
    dst_cell.value = value

    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format
    if src_cell.protection:
        dst_cell.protection = copy(src_cell.protection)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)
    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)


def _copy_row_dimension(ws, src_row: int, dst_row: int) -> None:
    """行高などの行設定をコピーする。"""
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    dst_dim.collapsed = src_dim.collapsed


def _template_merges(ws, template_start: int, block_rows: int) -> list[CellRange]:
    """テンプレートブロック内に完全に含まれる結合範囲を返す。"""
    template_end = template_start + block_rows - 1
    return [
        CellRange(str(merge))
        for merge in ws.merged_cells.ranges
        if template_start <= merge.min_row and merge.max_row <= template_end
    ]


def _copy_template_block(ws, template_start: int, dst_start: int,
                         block_rows: int, max_col: int,
                         merges: list[CellRange]) -> None:
    """直上ブロックの書式・数式・結合を追加ブロックへ複製する。"""
    row_offset = dst_start - template_start
    for row_offset_in_block in range(block_rows):
        src_row = template_start + row_offset_in_block
        dst_row = dst_start + row_offset_in_block
        _copy_row_dimension(ws, src_row, dst_row)
        for col in range(1, max_col + 1):
            _copy_cell(ws.cell(src_row, col), ws.cell(dst_row, col))

    existing_merges = {str(merge) for merge in ws.merged_cells.ranges}
    for merge in merges:
        shifted = CellRange(
            min_col=merge.min_col,
            min_row=merge.min_row + row_offset,
            max_col=merge.max_col,
            max_row=merge.max_row + row_offset,
        )
        if shifted.coord not in existing_merges:
            ws.merge_cells(shifted.coord)
            existing_merges.add(shifted.coord)


def _area_without_sheet(area: str) -> str:
    """印刷範囲文字列からシート名部分を除去する。"""
    if '!' in area:
        return area.split('!', 1)[1]
    return area


def _expand_print_area(ws, old_total_row: int, new_total_row: int,
                       logger: Logger) -> None:
    """行追加で合計行が下がった場合、既存の印刷範囲の下端も広げる。"""
    if not ws.print_area:
        return

    areas = ws.print_area if isinstance(ws.print_area, list) else [ws.print_area]
    expanded_areas: list[str] = []
    changed = False

    for area in areas:
        raw_area = _area_without_sheet(str(area))
        min_col, min_row, max_col, max_row = range_boundaries(raw_area)
        if max_row >= old_total_row:
            max_row = max(max_row, new_total_row)
            changed = True
        expanded_areas.append(
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    if changed:
        ws.print_area = expanded_areas
        logger.log(
            f"印刷範囲更新：合計行移動に合わせて{new_total_row}行目まで拡張")


def find_total_row(ws, settings: dict) -> int:
    """合計行キーワードをA列で検索して行番号を返す。見つからなければ0。"""
    keyword = settings['sum_keyword']
    for (cell,) in ws.iter_rows(min_col=1, max_col=1):
        if cell.value is not None and keyword in str(cell.value):
            return cell.row
    return 0


def clear_destination(ws, settings: dict) -> None:
    """転記先の名称・仕様・数量・単位・単価を値のみクリア（書式・数式は保持）"""
    total_row = find_total_row(ws, settings)
    start_row = settings['dst_start_row']

    if total_row > 0:
        end_row = total_row - 1
    else:
        end_row = 0
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                if cell.value is not None:
                    end_row = cell.row
        if end_row == 0:
            return

    for r in range(start_row, end_row + 1):
        _safe_set(ws, r, settings['dst_col_name'],  None)
        _safe_set(ws, r, settings['dst_col_spec'],  None)
        _safe_set(ws, r, settings['dst_col_qty'],   None)
        _safe_set(ws, r, settings['dst_col_unit'],  None)
        _safe_set(ws, r, settings['dst_col_price'], None)


def ensure_rows(ws, settings: dict, needed_blocks: int, logger: Logger) -> int:
    """
    必要なブロック数に対して行数が不足している場合、レイアウト行数単位で行を挿入する。
    戻り値：追加した行数（合計行が無い場合は0）
    """
    total_row = find_total_row(ws, settings)
    if total_row == 0:
        logger.log("合計行なし：行追加処理をスキップします")
        return 0

    start_row = settings['dst_start_row']
    block_rows = get_block_rows(settings)
    current_blocks = (total_row - start_row) // block_rows
    short_blocks = needed_blocks - current_blocks

    if short_blocks <= 0:
        return 0

    add_rows = short_blocks * block_rows
    old_total_row = total_row
    template_start = total_row - block_rows
    template_merges = _template_merges(ws, template_start, block_rows)
    max_col = max(ws.max_column, _col(settings['dst_col_amount']))

    for _ in range(short_blocks):
        total_row = find_total_row(ws, settings)
        ws.insert_rows(total_row, amount=block_rows)
        _copy_template_block(
            ws, template_start, total_row, block_rows, max_col, template_merges)
        # insert_rows は total_row の直前に挿入するため、
        # 合計行は total_row + block_rows に移動する（find_total_row で再取得）

    # 挿入された行の入力値をクリア。数式・書式・結合はテンプレートから残す。
    new_total_row = find_total_row(ws, settings)
    for r in range(new_total_row - add_rows, new_total_row):
        _safe_set(ws, r, settings['dst_col_name'],  None)
        _safe_set(ws, r, settings['dst_col_spec'],  None)
        _safe_set(ws, r, settings['dst_col_qty'],   None)
        _safe_set(ws, r, settings['dst_col_unit'],  None)
        _safe_set(ws, r, settings['dst_col_price'], None)

    _expand_print_area(ws, old_total_row, new_total_row, logger)

    return add_rows


def write_to_destination(ws, settings: dict, blocks: list[dict]) -> None:
    """選択されたレイアウトで明細ブロックを転記先に書き込む。"""
    layout = settings.get('dst_layout', LAYOUT_3_ROWS)
    if layout == LAYOUT_1_ROW:
        _write_1row_layout(ws, settings, blocks)
    elif layout == LAYOUT_2_ROWS:
        _write_2row_layout(ws, settings, blocks)
    else:
        _write_3row_layout(ws, settings, blocks)


def _write_common_values(ws, settings: dict, row: int, block: dict) -> None:
    """数量・単位・単価をブロック1行目へ書き込む。"""
    _safe_set(ws, row, settings['dst_col_qty'],   block['qty'])
    _safe_set(ws, row, settings['dst_col_unit'],  block['unit'])
    _safe_set(ws, row, settings['dst_col_price'], block['price'])


def _write_3row_layout(ws, settings: dict, blocks: list[dict]) -> None:
    """3行標準：現行帳票向けの配置。"""
    write_row = settings['dst_start_row']

    for block in blocks:
        # 行1：仕様1 / 数量 / 単位 / 単価（名称は空のまま）
        _safe_set(ws, write_row,     settings['dst_col_spec'],  block['spec1'])
        _write_common_values(ws, settings, write_row, block)
        # 行2：名称1 / 仕様2
        _safe_set(ws, write_row + 1, settings['dst_col_name'],  block['name1'])
        _safe_set(ws, write_row + 1, settings['dst_col_spec'],  block['spec2'])
        # 行3：名称2 / 仕様3
        _safe_set(ws, write_row + 2, settings['dst_col_name'],  block['name2'])
        _safe_set(ws, write_row + 2, settings['dst_col_spec'],  block['spec3'])

        write_row += 3


def _write_2row_layout(ws, settings: dict, blocks: list[dict]) -> None:
    """2行標準：上段に1行目情報と数量、下段に主名称と残り仕様を配置。"""
    write_row = settings['dst_start_row']

    for block in blocks:
        # 行1：名称1 / 仕様1 / 数量 / 単位 / 単価
        _safe_set(ws, write_row,     settings['dst_col_name'],  block['name1'])
        _safe_set(ws, write_row,     settings['dst_col_spec'],  block['spec1'])
        _write_common_values(ws, settings, write_row, block)
        # 行2：名称2 / 仕様2・仕様3
        _safe_set(ws, write_row + 1, settings['dst_col_name'],  block['name2'])
        _safe_set(ws, write_row + 1, settings['dst_col_spec'],
                  _join_lines(block['spec2'], block['spec3']))

        write_row += 2


def _write_1row_layout(ws, settings: dict, blocks: list[dict]) -> None:
    """1行標準：名称と仕様をセル内改行で1行に集約して配置。"""
    write_row = settings['dst_start_row']

    for block in blocks:
        _safe_set(ws, write_row, settings['dst_col_name'],
                  _join_lines(block['name1'], block['name2']))
        _safe_set(ws, write_row, settings['dst_col_spec'],
                  _join_lines(block['spec1'], block['spec2'], block['spec3']))
        _write_common_values(ws, settings, write_row, block)

        write_row += 1


def rebuild_sum_formula(ws, settings: dict, logger: Logger) -> str:
    """
    金額列のSUM式を完全再生成して合計行に書き込む。
    開始行からレイアウト行数おきにブロック1行目のセルを収集して =SUM(...) を組み立てる。
    """
    total_row = find_total_row(ws, settings)
    if total_row == 0:
        logger.log("合計行なし：SUM式更新をスキップします")
        return "(SUM式更新スキップ)"

    start_row  = settings['dst_start_row']
    block_rows = get_block_rows(settings)
    col_letter = settings['dst_col_amount'].upper()
    col_idx    = _col(settings['dst_col_amount'])

    cell_refs = [
        f"{col_letter}{r}"
        for r in range(start_row, total_row, block_rows)
    ]
    formula = f"=SUM({','.join(cell_refs)})"

    # 合計行の金額列セルに書き込む（結合セル対応）
    for merge in ws.merged_cells.ranges:
        if (merge.min_row <= total_row <= merge.max_row
                and merge.min_col <= col_idx <= merge.max_col):
            ws.cell(merge.min_row, merge.min_col).value = formula
            return formula

    ws.cell(total_row, col_idx).value = formula
    return formula
