"""
コピー元（内訳明細）Excelシートを読み取り、明細ブロックのリストを返す。

ブロック判定ルール：
  アンカー行  … 数量が数値 AND 単位に値あり
  上の行     … アンカー行の直前1〜2行（空白行はスキップ）で
               ・別のアンカーでない
               ・セクション見出しでない（「（内部）」等）
               ・前のアンカーの下の行でない（spec のみ・name なし）
               ・name または spec に値あり
  下の行     … アンカー行の直後1行で
               ・アンカーでない
               ・name が空 AND spec に値あり
               ・その次の行が次のアンカーでない（次ブロックの上の行と重複しない）
"""

from openpyxl.utils import column_index_from_string
from logger import Logger


def _col(letter: str) -> int:
    return column_index_from_string(letter.upper())


def _str(ws, row: int, col_letter: str) -> str:
    v = ws.cell(row, _col(col_letter)).value
    return str(v).strip() if v is not None else ''


def _val(ws, row: int, col_letter: str):
    return ws.cell(row, _col(col_letter)).value


def _get_last_row(ws, start_row: int) -> int:
    last = 0
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != '':
                last = cell.row
    return last


def _is_anchor(qty, unit: str) -> bool:
    if qty is None or qty == '':
        return False
    try:
        float(qty)
    except (ValueError, TypeError):
        return False
    return bool(unit)


def _is_blank(name: str, spec: str, qty, unit: str) -> bool:
    return not name and not spec and (qty is None or qty == '') and not unit


def _is_section_header(name: str, spec: str, qty, unit: str) -> bool:
    """「（内部）」「（外部）」等のセクション見出し行を判定"""
    if not name:
        return False
    has_parens = name.startswith('（') or name.startswith('(')
    no_data = not spec and (qty is None or qty == '') and not unit
    return has_parens and no_data


def read_source_data(ws, settings: dict, logger: Logger) -> list[dict]:
    start_row  = settings['src_start_row']
    col_name   = settings['src_col_name']
    col_spec   = settings['src_col_spec']
    col_qty    = settings['src_col_qty']
    col_unit   = settings['src_col_unit']
    col_price  = settings['src_col_price']

    last_row = _get_last_row(ws, start_row)
    if last_row == 0:
        return []

    # 全行をリストに展開（0-indexed: rows[0] = start_row）
    rows: list[dict] = []
    for r in range(start_row, last_row + 1):
        name  = _str(ws, r, col_name)
        spec  = _str(ws, r, col_spec)
        qty   = _val(ws, r, col_qty)
        unit  = _str(ws, r, col_unit)
        price = _val(ws, r, col_price)
        rows.append({
            'row': r, 'name': name, 'spec': spec,
            'qty': qty, 'unit': unit, 'price': price,
        })

    # パス1：アンカー行のインデックスを収集
    anchor_indices: set[int] = set()
    for i, row in enumerate(rows):
        if _is_anchor(row['qty'], row['unit']):
            anchor_indices.add(i)

    # パス2：各アンカーに対してブロックを構築
    blocks: list[dict] = []
    for i in sorted(anchor_indices):
        anchor = rows[i]

        # ── 上の行を探す（最大2行遡る、空白行はスキップ）──────────────
        upper = None
        for back in [1, 2]:
            j = i - back
            if j < 0:
                break
            cand = rows[j]

            if _is_blank(cand['name'], cand['spec'], cand['qty'], cand['unit']):
                continue  # 空白行はスキップして次の back を試す

            # 非空白行に到達 → この行で判定を確定する
            if j in anchor_indices:
                break  # 別のアンカーに当たった

            if _is_section_header(cand['name'], cand['spec'], cand['qty'], cand['unit']):
                break  # セクション見出しは上の行として使わない

            # 直前がアンカーの場合、この行は前ブロックの「下の行」なので使わない
            if (not cand['name'] and cand['spec']
                    and (j - 1) >= 0 and (j - 1) in anchor_indices):
                break

            if cand['name'] or cand['spec']:
                upper = cand
            break  # 非空白行が見つかった時点でループ終了

        # ── 下の行を探す（直後1行のみ）────────────────────────────────
        lower_spec = ''
        k = i + 1
        if k < len(rows):
            cand = rows[k]
            if (not _is_anchor(cand['qty'], cand['unit'])
                    and not cand['name']
                    and cand['spec']):
                # さらに次の行が次のアンカーでなければ採用
                if k + 1 >= len(rows) or (k + 1) not in anchor_indices:
                    lower_spec = cand['spec']
                # 4行目以降に仕様データがあれば警告
                if lower_spec and k + 1 < len(rows):
                    nxt = rows[k + 1]
                    if (not _is_anchor(nxt['qty'], nxt['unit'])
                            and not nxt['name'] and nxt['spec']
                            and (k + 1) not in anchor_indices):
                        logger.log(
                            f"警告：行{anchor['row']}の明細は仕様が4行以上あります。"
                            "3行目以降は省略されました。"
                        )

        blocks.append({
            'name1': upper['name'] if upper else '',
            'name2': anchor['name'],
            'spec1': upper['spec'] if upper else '',
            'spec2': anchor['spec'],
            'spec3': lower_spec,
            'qty':   anchor['qty'],
            'unit':  anchor['unit'],
            'price': anchor['price'],
        })

    return blocks
