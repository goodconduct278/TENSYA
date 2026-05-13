"""
転記ツール - メイン（GUI）

起動方法：
    python main.py
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl

from backup import backup_file
from logger import Logger
from reader import read_source_data
from settings_manager import load_settings, save_settings
from writer import (
    clear_destination,
    ensure_rows,
    rebuild_sum_formula,
    write_to_destination,
)

BACKUP_MODES = [
    '常に作成',
    '日付フォルダにまとめる',
    '直前1件のみ保持',
    '実行時に確認',
    '作成しない',
]

DST_LAYOUTS = [
    '3行標準',
    '2行標準',
    '1行標準',
]


# ──────────────────────────────────────────────────────────────────────────────
# 設定ダイアログ
# ──────────────────────────────────────────────────────────────────────────────

class SettingsDialog(tk.Toplevel):
    _FIELDS = [
        ('src_sheet',      'コピー元シート名'),
        ('src_start_row',  'コピー元開始行'),
        ('src_col_name',   'コピー元名称列'),
        ('src_col_spec',   'コピー元仕様列'),
        ('src_col_qty',    'コピー元数量列'),
        ('src_col_unit',   'コピー元単位列'),
        ('src_col_price',  'コピー元単価列'),
        ('dst_sheet',      '転記先シート名'),
        ('dst_start_row',  '転記先開始行'),
        ('dst_col_name',   '転記先名称列'),
        ('dst_col_spec',   '転記先仕様列'),
        ('dst_col_qty',    '転記先数量列'),
        ('dst_col_unit',   '転記先単位列'),
        ('dst_col_price',  '転記先単価列'),
        ('dst_col_amount', '転記先金額列'),
        ('dst_layout',     '転記先レイアウト'),
        ('sum_keyword',    '合計行キーワード'),
    ]
    _INT_KEYS = {'src_start_row', 'dst_start_row'}
    _CHOICE_FIELDS = {'dst_layout': DST_LAYOUTS}

    def __init__(self, parent: tk.Tk, settings: dict):
        super().__init__(parent)
        self.title('設定')
        self.resizable(False, False)
        self.grab_set()
        self.settings = settings
        self._vars: dict[str, tk.StringVar] = {}
        self._build()

    def _build(self) -> None:
        pad = {'padx': 8, 'pady': 3}
        for i, (key, label) in enumerate(self._FIELDS):
            tk.Label(self, text=label, anchor='w', width=18).grid(
                row=i, column=0, **pad, sticky='w')
            var = tk.StringVar(value=str(self.settings.get(key, '')))
            self._vars[key] = var
            if key in self._CHOICE_FIELDS:
                ttk.Combobox(
                    self, textvariable=var, values=self._CHOICE_FIELDS[key],
                    state='readonly', width=20,
                ).grid(row=i, column=1, **pad)
            else:
                tk.Entry(self, textvariable=var, width=22).grid(
                    row=i, column=1, **pad)

        n = len(self._FIELDS)
        frame = tk.Frame(self)
        frame.grid(row=n, column=0, columnspan=2, pady=8)
        tk.Button(frame, text='保存して閉じる', command=self._save).pack(
            side='left', padx=6)
        tk.Button(frame, text='キャンセル', command=self.destroy).pack(
            side='left', padx=6)

    def _save(self) -> None:
        for key, var in self._vars.items():
            raw = var.get().strip()
            if key in self._INT_KEYS:
                try:
                    self.settings[key] = int(raw)
                except ValueError:
                    messagebox.showwarning(
                        '設定エラー', f'「{key}」は整数で入力してください',
                        parent=self)
                    return
            else:
                self.settings[key] = raw
        save_settings(self.settings)
        messagebox.showinfo('設定', '設定を保存しました', parent=self)
        self.destroy()


# ──────────────────────────────────────────────────────────────────────────────
# メインウィンドウ
# ──────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('転記ツール')
        self.resizable(False, False)
        self.settings = load_settings()
        self._build()

    def _build(self) -> None:
        pad = {'padx': 10, 'pady': 5}

        # コピー元
        tk.Label(self, text='コピー元：', anchor='w').grid(
            row=0, column=0, **pad, sticky='w')
        self._src_var = tk.StringVar()
        tk.Entry(self, textvariable=self._src_var, width=60,
                 state='readonly').grid(row=0, column=1, **pad)
        tk.Button(self, text='選択', width=6,
                  command=self._select_src).grid(row=0, column=2, **pad)

        # 転記先
        tk.Label(self, text='転記先：', anchor='w').grid(
            row=1, column=0, **pad, sticky='w')
        self._dst_var = tk.StringVar()
        tk.Entry(self, textvariable=self._dst_var, width=60,
                 state='readonly').grid(row=1, column=1, **pad)
        tk.Button(self, text='選択', width=6,
                  command=self._select_dst).grid(row=1, column=2, **pad)

        # バックアップモード
        tk.Label(self, text='バックアップ：', anchor='w').grid(
            row=2, column=0, **pad, sticky='w')
        self._backup_var = tk.StringVar(
            value=self.settings.get('backup_mode', '常に作成'))
        ttk.Combobox(
            self, textvariable=self._backup_var,
            values=BACKUP_MODES, state='readonly', width=20,
        ).grid(row=2, column=1, **pad, sticky='w')

        # ボタン行
        btn_frame = tk.Frame(self)
        btn_frame.grid(row=3, column=0, columnspan=3, pady=10)
        tk.Button(
            btn_frame, text='転記実行', width=16,
            bg='#4CAF50', fg='white', font=('', 11, 'bold'),
            command=self._run,
        ).pack(side='left', padx=8)
        tk.Button(
            btn_frame, text='設定', width=8,
            command=self._open_settings,
        ).pack(side='left', padx=8)

        # ステータスバー
        self._status_var = tk.StringVar(value='ファイルを選択してください')
        tk.Label(self, textvariable=self._status_var, fg='gray',
                 anchor='w').grid(row=4, column=0, columnspan=3,
                                  padx=10, pady=(0, 8), sticky='w')

    # ── ファイル選択 ─────────────────────────────────────────────────────────

    def _select_src(self) -> None:
        path = filedialog.askopenfilename(
            title='コピー元（内訳明細）ファイルを選択',
            filetypes=[('Excelファイル', '*.xlsx *.xlsm *.xls')])
        if path:
            self._src_var.set(path)
            self._status_var.set('コピー元を選択しました')

    def _select_dst(self) -> None:
        path = filedialog.askopenfilename(
            title='転記先（見積書）ファイルを選択',
            filetypes=[('Excelファイル', '*.xlsx *.xlsm *.xls')])
        if path:
            self._dst_var.set(path)
            self._status_var.set('転記先を選択しました')

    # ── 設定ダイアログ ────────────────────────────────────────────────────────

    def _open_settings(self) -> None:
        SettingsDialog(self, self.settings)

    # ── 転記実行 ──────────────────────────────────────────────────────────────

    def _run(self) -> None:
        src_path = self._src_var.get()
        dst_path = self._dst_var.get()

        if not src_path:
            messagebox.showwarning('転記ツール', 'コピー元ファイルを選択してください')
            return
        if not dst_path:
            messagebox.showwarning('転記ツール', '転記先ファイルを選択してください')
            return

        # バックアップモードを設定に反映・保存
        self.settings['backup_mode'] = self._backup_var.get()
        save_settings(self.settings)

        logger = Logger()
        logger.log('===== 転記処理開始 =====')
        logger.log(f'コピー元：{src_path}')
        logger.log(f'転記先　：{dst_path}')

        try:
            self._execute(src_path, dst_path, logger)
        finally:
            logger.flush(dst_path)

    def _execute(self, src_path: str, dst_path: str, logger: Logger) -> None:
        mode = self.settings['backup_mode']

        # 「実行時に確認」モードのダイアログ
        if mode == '実行時に確認':
            if not messagebox.askyesno('バックアップ確認', 'バックアップを作成しますか？'):
                mode = '作成しない'

        # バックアップ
        if mode != '作成しない':
            self._status_var.set('バックアップ作成中...')
            self.update()
            if not backup_file(dst_path, mode, logger):
                messagebox.showerror(
                    '転記ツール',
                    'バックアップに失敗しました。転記を中止します。\n'
                    '転記先ファイルが別のアプリで開かれていないか確認してください。')
                self._status_var.set('バックアップ失敗')
                return

        # コピー元を読み取り
        self._status_var.set('コピー元を読み取り中...')
        self.update()

        try:
            src_wb = openpyxl.load_workbook(
                src_path, read_only=True, data_only=True)
        except Exception as e:
            messagebox.showerror('転記ツール', f'コピー元ファイルを開けません：\n{e}')
            logger.log(f'エラー：コピー元を開けません {e}')
            return

        try:
            src_sheet = self.settings['src_sheet']
            if src_sheet not in src_wb.sheetnames:
                available = '、'.join(src_wb.sheetnames)
                messagebox.showerror(
                    '転記ツール',
                    f'コピー元シートが見つかりません：[{src_sheet}]\n\n'
                    f'利用可能なシート：{available}')
                logger.log(f'エラー：シート不在 [{src_sheet}]')
                return

            try:
                blocks = read_source_data(src_wb[src_sheet], self.settings, logger)
            except Exception as e:
                messagebox.showerror(
                    '転記ツール', f'コピー元の読み取りに失敗しました：\n{e}')
                logger.log(f'エラー：コピー元読み取り失敗 {e}')
                self._status_var.set('コピー元読み取りエラー')
                return
        finally:
            src_wb.close()

        logger.log(f'読み取り件数：{len(blocks)}件')

        if not blocks:
            messagebox.showwarning('転記ツール', '転記できる明細がありませんでした。\n'
                                   '設定のシート名・列・開始行を確認してください。')
            self._status_var.set('明細が見つかりませんでした')
            return

        # 転記先を開いて書き込み
        self._status_var.set('転記中...')
        self.update()

        try:
            dst_wb = openpyxl.load_workbook(dst_path)
        except Exception as e:
            messagebox.showerror('転記ツール', f'転記先ファイルを開けません：\n{e}')
            logger.log(f'エラー：転記先を開けません {e}')
            return

        try:
            dst_sheet = self.settings['dst_sheet']
            if dst_sheet not in dst_wb.sheetnames:
                available = '、'.join(dst_wb.sheetnames)
                messagebox.showerror(
                    '転記ツール',
                    f'転記先シートが見つかりません：[{dst_sheet}]\n\n'
                    f'利用可能なシート：{available}')
                logger.log(f'エラー：シート不在 [{dst_sheet}]')
                return

            dst_ws = dst_wb[dst_sheet]

            clear_destination(dst_ws, self.settings)
            logger.log('既存データクリア完了')

            added = ensure_rows(dst_ws, self.settings, len(blocks), logger)
            if added > 0:
                logger.log(f'行追加：{added}行')
            else:
                logger.log('行追加：なし')

            write_to_destination(dst_ws, self.settings, blocks)
            logger.log(f'転記件数：{len(blocks)}件')

            formula = rebuild_sum_formula(dst_ws, self.settings, logger)
            logger.log(f'更新後SUM式：{formula}')

            dst_wb.save(dst_path)
            logger.log('転記先ファイル保存完了')
            logger.log('===== 転記処理正常終了 =====')

            self._status_var.set(f'完了：{len(blocks)}件転記しました')
            messagebox.showinfo('転記完了', f'{len(blocks)}件の転記が完了しました。')

        except Exception as e:
            logger.log(f'エラー：{e}')
            messagebox.showerror('転記ツール', f'転記中にエラーが発生しました：\n{e}')
            self._status_var.set('エラーが発生しました')
        finally:
            dst_wb.close()


# ──────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app = App()
    app.mainloop()
